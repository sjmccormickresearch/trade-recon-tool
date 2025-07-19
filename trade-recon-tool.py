import sqlite3
import pandas as pd
import os
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# Paths
db_path = os.path.join(os.path.dirname(__file__), 'trades.db')
output_path = os.path.join(os.path.dirname(__file__), 'reconciliation_report.xlsx')

# Connect to database
conn = sqlite3.connect(db_path)

# Load both tables
internal = pd.read_sql_query("SELECT * FROM internal_trades", conn)
broker = pd.read_sql_query("SELECT * FROM broker_trades", conn)

# Merge on TradeID
merged = pd.merge(internal, broker, on='TradeID', how='outer', suffixes=('_internal', '_broker'))

# Flag mismatches
def check_row(row):
    if pd.isna(row['Ticker_internal']):
        return 'Missing Internally'
    elif pd.isna(row['Ticker_broker']):
        return 'Missing at Broker'
    elif row['Quantity_internal'] != row['Quantity_broker'] or row['Price_internal'] != row['Price_broker']:
        return 'Mismatch'
    else:
        return 'Match'

merged['Status'] = merged.apply(check_row, axis=1)

# Export to Excel
merged.to_excel(output_path, index=False)

# Add color formatting
wb = load_workbook(output_path)
ws = wb.active

fill_match = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")     # Green
fill_mismatch = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")  # Red
fill_missing = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")   # Yellow

for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
    status_cell = row[-1]  # last column is Status
    if status_cell.value == "Match":
        for cell in row:
            cell.fill = fill_match
    elif status_cell.value == "Mismatch":
        for cell in row:
            cell.fill = fill_mismatch
    else:  # Missing Internally or Missing at Broker
        for cell in row:
            cell.fill = fill_missing

wb.save(output_path)
print(f"✅ Reconciliation report saved to: {output_path}")
